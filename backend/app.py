import base64
import hashlib
import io
import json
import os
import re
import time
import traceback
import threading
import uuid
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, make_response, send_file
from flask_cors import CORS
from openai import OpenAI
from dotenv import load_dotenv
from puzzle_engine import PuzzleEngine, PuzzleError
from analytics_store import (
    AnalyticsStore,
    normalize_client_id,
    parse_completion_time_to_seconds,
)
try:
    from PIL import Image, ImageOps
except Exception:
    Image = None
    ImageOps = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception:
    Workbook = None

# 加载环境变量
load_dotenv()

# 初始化Flask应用
app = Flask(__name__)
CORS(app)  # 允许跨域请求

# 拼图引擎（服务端权威状态）
puzzle_engine = PuzzleEngine()

# 行为分析存储（SQLite）
analytics_store = AnalyticsStore(
    db_path=os.path.join(os.path.dirname(__file__), "data", "behavior_analytics.db")
)

# 延迟初始化客户端
client = None
bailian_client = None
IMAGE_VALIDATION_CACHE = {}

# 管理员密码
ADMIN_PASSWORD = "123456"

# 异步报告生成任务存储
report_generation_tasks = {}
report_generation_lock = threading.Lock()

def get_client():
    """获取或创建OpenAI客户端"""
    global client
    if client is None:
        client = OpenAI(
            api_key=os.environ.get("DEEPSEEK_API_KEY"),
            base_url="https://api.deepseek.com"
        )
    return client


def get_bailian_client():
    """获取或创建阿里云百炼（DashScope 兼容模式）客户端。"""
    global bailian_client
    if bailian_client is None:
        api_key = os.environ.get("BAILIAN_API_KEY") or os.environ.get("DASHSCOPE_API_KEY")
        if not api_key:
            raise RuntimeError("未配置 BAILIAN_API_KEY（或 DASHSCOPE_API_KEY）")
        bailian_client = OpenAI(
            api_key=api_key,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
        )
    return bailian_client

SYSTEM_PROMPT = """
# 角色定位
你是高校心理健康教育场景中的心理学科普助手，服务对象是大学生。
你的工作是基于房树人主题拼图行为数据，给出温和、可执行、非诊断的心理观察与自我觉察建议。

# 核心任务
我会给你“本次拼图行为数据”和“近期行为数据摘要”。
请你输出一份面向大学生的心理分析报告，覆盖学业任务、求职实习、人际协作、作息管理等校园常见压力情境。

# 证据使用规则
1. 以“本次拼图行为数据”为主证据，“近期行为数据摘要”为辅证据。
2. 当本次数据与近期趋势冲突时，优先解释本次数据，并把历史信息写成“可能的背景趋势”。
3. 当近期样本偏少时，明确提示“样本有限”，禁止下绝对结论。
4. 每条判断都要映射到行为证据（用时、步数、间隔、修改次数、顺序、回退等）。

# 安全与表达限制
1. 禁止医学诊断、病理标签或治疗结论。
2. 禁止“你有问题”“你患有”等负向定性表达。
3. 只能做心理科普和自我觉察引导，不替代专业诊疗。
4. 禁止使用表情符号和无关特殊符号，中文表达要清晰整洁。
5. 仅分析两类图像：系统内置图像，或已通过“房+树+人三要素”校验的用户上传图像。

# 输出结构（必须按以下标题输出，不要加数字序号）
## 总体状态观察
## 房树人维度解读
### 房屋
### 树木
### 人物
## 校园情境关联
## 可执行建议
### 一句话心理提示（120字以内）

# 结尾要求
在报告最后另起一行，补充一句温暖鼓励语（不超过30字）。
"""

REWRITE_PROMPT = """
你正在修订一份心理分析报告。请保留事实证据，不扩写虚构信息，仅修正格式与措辞使其合规。
必须满足：
1. 面向大学生场景；
2. 不使用医学诊断/病理术语；
3. 标题使用固定 Markdown 结构，不加数字序号；
4. “一句话心理提示”不超过120字；
5. 报告最后单独一行给出温暖鼓励语（不超过30字）；
6. 不使用表情符号和无关特殊符号。
"""

FORBIDDEN_TERMS = [
    "抑郁症",
    "焦虑症",
    "精神分裂",
    "双相障碍",
    "强迫症",
    "人格障碍",
    "确诊",
    "诊断为",
    "患有",
    "病理",
    "治疗方案",
    "处方",
    "药物治疗",
]

# 允许分析的图片列表（只有这些图片可以被分析）
ALLOWED_IMAGES = [
    'photo/1.png',
    'photo/2.png',
    'photo/3.jpg',
    'photo/4.jpg'
]
FRONTEND_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend"))
IMAGE_VALIDATION_CACHE_TTL_SECONDS = 12 * 3600
DEFAULT_BAILIAN_VISION_MODELS = [
    "qwen-vl-max-latest",
    "qwen-vl-max",
    "qwen2.5-vl-72b-instruct",
]


def _load_custom_image_limit_mb():
    """加载用户上传图片大小上限（MB），默认20MB，范围1~50MB。"""
    raw = (os.environ.get("CUSTOM_IMAGE_MAX_MB") or "").strip()
    default_mb = 20.0
    if not raw:
        return default_mb
    try:
        parsed = float(raw)
    except ValueError:
        return default_mb
    return max(1.0, min(50.0, parsed))


def _load_bailian_image_limit_mb():
    """加载发送给百炼多模态接口的图片大小上限（MB）。"""
    raw = (os.environ.get("BAILIAN_IMAGE_MAX_MB") or "").strip()
    default_mb = 9.0
    if not raw:
        return default_mb
    try:
        parsed = float(raw)
    except ValueError:
        return default_mb
    return max(1.0, min(20.0, parsed))


CUSTOM_IMAGE_MAX_MB = _load_custom_image_limit_mb()
CUSTOM_IMAGE_MAX_BYTES = int(CUSTOM_IMAGE_MAX_MB * 1024 * 1024)
BAILIAN_IMAGE_MAX_MB = _load_bailian_image_limit_mb()
BAILIAN_IMAGE_MAX_BYTES = int(BAILIAN_IMAGE_MAX_MB * 1024 * 1024)
BAILIAN_IMAGE_MAX_SIDE = 4096


def _is_builtin_image(image_source):
    return isinstance(image_source, str) and image_source in ALLOWED_IMAGES


def _is_data_image(image_source):
    return isinstance(image_source, str) and image_source.startswith("data:image/")


def _extract_data_image_payload(image_source):
    if not _is_data_image(image_source):
        raise PuzzleError("仅支持 image/* 数据URL")
    head, sep, payload = image_source.partition(",")
    if not sep:
        raise PuzzleError("图片数据格式不正确")
    mime_match = re.match(r"^data:(image\/[a-zA-Z0-9.+-]+)(;.*)?$", head, flags=re.IGNORECASE)
    if not mime_match:
        raise PuzzleError("图片MIME类型不合法")
    params = [p.strip().lower() for p in head.split(";")[1:]]
    if "base64" not in params:
        raise PuzzleError("仅支持 base64 编码的图片数据")
    try:
        raw = base64.b64decode(payload, validate=True)
    except Exception as e:
        raise PuzzleError(f"图片base64解析失败: {e}") from e
    if not raw:
        raise PuzzleError("图片内容为空")
    if len(raw) > CUSTOM_IMAGE_MAX_BYTES:
        raise PuzzleError(f"图片过大，最大支持 {CUSTOM_IMAGE_MAX_MB:g}MB")
    return mime_match.group(1).lower(), raw


def _prepare_image_for_bailian(mime_type, raw):
    """将上传图片压缩到百炼多模态可接受范围。"""
    if len(raw) <= BAILIAN_IMAGE_MAX_BYTES:
        return mime_type, raw

    if Image is None:
        raise PuzzleError(
            f"图片大于 {BAILIAN_IMAGE_MAX_MB:g}MB，且服务端缺少压缩组件。请先压缩后重试。"
        )

    try:
        with Image.open(io.BytesIO(raw)) as img:
            if ImageOps is not None:
                img = ImageOps.exif_transpose(img)
            img.load()

            if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
                rgba = img.convert("RGBA")
                base = Image.new("RGB", rgba.size, (255, 255, 255))
                base.paste(rgba, mask=rgba.split()[-1])
                img = base
            else:
                img = img.convert("RGB")

            # 优先限制最大边，避免超大分辨率直接触发模型侧拒绝。
            if max(img.size) > BAILIAN_IMAGE_MAX_SIDE:
                resized = img.copy()
                resized.thumbnail((BAILIAN_IMAGE_MAX_SIDE, BAILIAN_IMAGE_MAX_SIDE), Image.Resampling.LANCZOS)
                img = resized

            quality_levels = [90, 85, 80, 75, 70, 65, 60, 55]
            side_limits = [4096, 3072, 2560, 2048, 1600, 1280]
            best_bytes = None

            for side in side_limits:
                work = img.copy()
                if max(work.size) > side:
                    work.thumbnail((side, side), Image.Resampling.LANCZOS)

                for quality in quality_levels:
                    buff = io.BytesIO()
                    work.save(
                        buff,
                        format="JPEG",
                        optimize=True,
                        progressive=True,
                        quality=quality
                    )
                    data = buff.getvalue()
                    if best_bytes is None or len(data) < len(best_bytes):
                        best_bytes = data
                    if len(data) <= BAILIAN_IMAGE_MAX_BYTES:
                        return "image/jpeg", data

            if best_bytes and len(best_bytes) <= BAILIAN_IMAGE_MAX_BYTES:
                return "image/jpeg", best_bytes

    except PuzzleError:
        raise
    except Exception as e:
        raise PuzzleError(f"图片压缩失败：{e}") from e

    raise PuzzleError(f"图片体积仍过大，请压缩到 {BAILIAN_IMAGE_MAX_MB:g}MB 以内后重试")


def _compact_image_source(image_source, image_check=None):
    if _is_builtin_image(image_source):
        return image_source
    if isinstance(image_check, dict) and image_check.get("imageId"):
        return f"custom/{image_check['imageId']}"
    if _is_data_image(image_source):
        return "custom/uploaded"
    return str(image_source or "")[:80]


def _safe_image_prompt_ref(image_source, image_check=None):
    if _is_builtin_image(image_source):
        return image_source
    if isinstance(image_check, dict) and image_check.get("imageId"):
        return f"用户上传图片（ID: {image_check['imageId']}）"
    return "用户上传图片（已通过内容校验）"


def _cleanup_image_validation_cache():
    now = time.time()
    expired = [
        key for key, val in IMAGE_VALIDATION_CACHE.items()
        if now - float(val.get("ts", 0)) > IMAGE_VALIDATION_CACHE_TTL_SECONDS
    ]
    for key in expired:
        IMAGE_VALIDATION_CACHE.pop(key, None)


def _get_cached_image_validation(image_hash):
    _cleanup_image_validation_cache()
    hit = IMAGE_VALIDATION_CACHE.get(image_hash)
    if not hit:
        return None
    return dict(hit.get("result", {}))


def _set_cached_image_validation(image_hash, result):
    IMAGE_VALIDATION_CACHE[image_hash] = {
        "ts": time.time(),
        "result": dict(result),
    }


def _parse_json_from_text(text):
    raw = str(text or "").strip()
    if not raw:
        raise ValueError("模型返回为空")
    # 优先直接解析
    try:
        return json.loads(raw)
    except Exception:
        pass
    # 尝试提取JSON代码块
    fenced = re.findall(r"```(?:json)?\s*(\{.*?\})\s*```", raw, flags=re.S)
    for chunk in fenced:
        try:
            return json.loads(chunk)
        except Exception:
            continue
    # 尝试提取首个对象
    match = re.search(r"\{.*\}", raw, flags=re.S)
    if match:
        return json.loads(match.group(0))
    raise ValueError("无法从模型输出中解析JSON")


def _to_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    txt = str(value or "").strip().lower()
    return txt in {"true", "1", "yes", "y", "是", "有", "包含", "存在"}


def _get_bailian_vision_models():
    custom = (os.environ.get("BAILIAN_VISION_MODEL") or "").strip()
    if custom:
        models = [m.strip() for m in custom.split(",") if m.strip()]
        if models:
            return models
    return DEFAULT_BAILIAN_VISION_MODELS


def _check_custom_image_with_bailian(image_source):
    """调用多模态模型检查图片是否同时包含房子、树、人物。"""
    mime_type, raw = _extract_data_image_payload(image_source)
    image_hash = hashlib.sha256(raw).hexdigest()
    cached = _get_cached_image_validation(image_hash)
    if cached:
        return cached

    prepared_mime, prepared_raw = _prepare_image_for_bailian(mime_type, raw)
    prepared_source = (
        f"data:{prepared_mime};base64,{base64.b64encode(prepared_raw).decode('ascii')}"
    )

    client_mm = get_bailian_client()
    last_err = None
    parsed = None

    for model_name in _get_bailian_vision_models():
        try:
            response = client_mm.chat.completions.create(
                model=model_name,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "你是图片要素审核器。任务：判断图片中是否同时存在“房子、树、人物”三种元素。"
                            "只输出JSON，不要任何解释性文字。"
                        ),
                    },
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": (
                                    "请判断图片是否包含以下元素：house(房子)、tree(树)、person(人物)。"
                                    "输出JSON格式："
                                    "{\"house\":true/false,\"tree\":true/false,\"person\":true/false,"
                                    "\"all_present\":true/false,\"reason\":\"不超过40字\"}"
                                ),
                            },
                            {
                                "type": "image_url",
                                "image_url": {"url": prepared_source},
                            },
                        ],
                    },
                ],
                temperature=0,
                max_tokens=180,
                stream=False,
            )
            content = response.choices[0].message.content
            parsed = _parse_json_from_text(content)
            if isinstance(parsed, dict):
                parsed["_model"] = model_name
                break
        except Exception as e:
            last_err = e
            continue

    if not isinstance(parsed, dict):
        if last_err:
            raise RuntimeError(f"图片要素校验失败: {last_err}")
        raise RuntimeError("图片要素校验失败: 模型未返回可解析结果")

    house = _to_bool(parsed.get("house"))
    tree = _to_bool(parsed.get("tree"))
    person = _to_bool(parsed.get("person"))
    all_present = _to_bool(parsed.get("all_present")) and house and tree and person
    reason = str(parsed.get("reason", "") or "").strip()[:120]

    result = {
        "valid": all_present,
        "isCustom": True,
        "elements": {"house": house, "tree": tree, "person": person},
        "allPresent": all_present,
        "reason": reason,
        "imageId": image_hash[:16],
        "model": parsed.get("_model", ""),
        "message": "图片校验通过，已检测到房子、树、人物三种元素。" if all_present else "图片缺少房子、树、人物中的至少一种元素。",
    }
    _set_cached_image_validation(image_hash, result)
    return result


def _validate_image_source_rule(image_source):
    """统一的图片规则校验：内置图片直接通过；自定义图片必须含房树人三元素。"""
    if _is_builtin_image(image_source):
        return {
            "valid": True,
            "isCustom": False,
            "allPresent": True,
            "elements": {"house": True, "tree": True, "person": True},
            "imageId": "",
            "message": "内置房树人图像，可直接使用。",
        }

    if str(image_source or "").startswith("blob:"):
        return {
            "valid": False,
            "isCustom": True,
            "allPresent": False,
            "elements": {"house": False, "tree": False, "person": False},
            "imageId": "",
            "message": "浏览器临时blob地址无法在服务端校验，请改为文件上传后生成的图片数据。",
        }

    if _is_data_image(image_source):
        try:
            return _check_custom_image_with_bailian(image_source)
        except Exception as e:
            return {
                "valid": False,
                "isCustom": True,
                "allPresent": False,
                "elements": {"house": False, "tree": False, "person": False},
                "imageId": "",
                "message": f"图片校验服务暂不可用：{str(e)[:120]}",
            }

    return {
        "valid": False,
        "isCustom": False,
        "allPresent": False,
        "elements": {"house": False, "tree": False, "person": False},
        "imageId": "",
        "message": "图片来源不合法。请选择内置图片，或上传包含房子、树、人物三要素的图片。",
    }


def _normalize_intervals(value):
    if isinstance(value, list):
        return [float(x) for x in value if isinstance(x, (int, float))]
    if isinstance(value, str):
        return [float(x) for x in re.findall(r"\d+(?:\.\d+)?", value)]
    return []


def _build_behavior_tags(game_data):
    tags = []
    if not isinstance(game_data, dict):
        return tags

    completion_seconds = parse_completion_time_to_seconds(game_data.get("completionTime"))
    move_count = int(game_data.get("moveCount", 0) or 0)
    modification_count = int(game_data.get("modificationCount", 0) or 0)
    grid_size = int(game_data.get("gridSize", 0) or 0)
    intervals = _normalize_intervals(game_data.get("timeIntervals"))
    avg_interval = round(sum(intervals) / len(intervals), 2) if intervals else None
    baseline_moves = max(1, grid_size * grid_size) if grid_size > 0 else 9

    if completion_seconds is not None and completion_seconds <= 120 and move_count <= baseline_moves * 2:
        tags.append("执行节奏较快，目标感明确")
    elif completion_seconds is not None and completion_seconds >= 360:
        tags.append("完成耗时较长，可能存在任务负荷或分心压力")

    if avg_interval is not None and avg_interval >= 4.0:
        tags.append("关键步骤停顿较多，决策更谨慎")
    elif avg_interval is not None and avg_interval <= 1.6:
        tags.append("动作间隔较短，过程较流畅")

    if modification_count >= max(6, baseline_moves // 2):
        tags.append("反复调整较多，可能自我要求偏高")
    elif modification_count == 0:
        tags.append("几乎无二次修改，当前判断较果断")

    if not tags:
        tags.append("整体节奏平稳，表现出持续推进任务的能力")
    return tags


def _build_fallback_brief_hint(game_data):
    tags = _build_behavior_tags(game_data)
    if any("停顿较多" in tag for tag in tags):
        hint = "你在关键步骤更谨慎，可能近期学业或人际压力偏高，先给自己一点缓冲再行动会更稳。"
    elif any("反复调整" in tag for tag in tags):
        hint = "你愿意反复调整到更满意，体现了责任感，也建议适当降低完美要求，给自己留出松弛空间。"
    elif any("节奏较快" in tag or "流畅" in tag for tag in tags):
        hint = "你在这次任务里展现了清晰目标感和执行力，面对学习安排时也具备较好的掌控感。"
    else:
        hint = "你的完成过程总体稳定，说明你具备持续推进任务的能力，遇到压力时保持当下节奏会更有帮助。"
    compact = re.sub(r"\s+", "", hint)
    return compact[:120].rstrip("，,。;；:：") + "。"


def _sanitize_report_line(text):
    cleaned = str(text or "")
    cleaned = re.sub(r"[\x00-\x1f\x7f-\x9f]", " ", cleaned)
    cleaned = re.sub(r"[\U0001F000-\U0001FAFF\U00002600-\U000027BF]", " ", cleaned)
    cleaned = re.sub(r"[★☆◆◇■□●○▶▷►▪▫※◎▲△▽▼•◦]", " ", cleaned)
    cleaned = re.sub(r"[|]+", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip()


def _clean_report_markdown(report):
    lines = str(report or "").replace("\r\n", "\n").replace("\r", "\n").split("\n")
    out = []
    for raw in lines:
        line = raw.strip()
        if not line:
            if out and out[-1] != "":
                out.append("")
            continue

        heading_match = re.match(r"^(#{1,3})\s*(.+)$", line)
        if heading_match:
            level, title = heading_match.groups()
            title = re.sub(r"^(?:\d+\s*[.)、]?\s*)+", "", title).strip()
            title = _sanitize_report_line(title)
            if title:
                out.append(f"{level} {title}")
            continue

        line = re.sub(r"^(?:\d+\s*[.)、]?\s*){2,}", "", line).strip()
        line = _sanitize_report_line(line)
        if line:
            out.append(line)

    while out and out[-1] == "":
        out.pop()
    return "\n".join(out).strip()


def _extract_brief_hint(report):
    lines = str(report or "").splitlines()
    capture = False
    for raw in lines:
        line = raw.strip()
        if not line:
            continue
        if line.startswith("#"):
            heading = re.sub(r"^#{1,3}\s*", "", line).strip()
            if "一句话心理提示" in heading:
                capture = True
                continue
            if capture:
                break
        elif capture:
            return _sanitize_report_line(line)
    return ""


def _count_compact_chars(text):
    return len(re.sub(r"\s+", "", str(text or "")))


def _has_encouraging_tail(report):
    lines = [ln.strip() for ln in str(report or "").splitlines() if ln.strip()]
    if not lines:
        return False
    tail = lines[-1]
    if tail.startswith("#"):
        return False
    keywords = ["加油", "继续", "你可以", "值得", "很好", "相信", "已经在", "勇敢", "稳步"]
    return any(k in tail for k in keywords)


def _collect_report_issues(report):
    issues = []
    text = str(report or "")
    if not text.strip():
        issues.append("报告为空")
        return issues

    hit_terms = [term for term in FORBIDDEN_TERMS if term in text]
    if hit_terms:
        issues.append("出现诊断或病理词汇: " + "、".join(sorted(set(hit_terms))))

    if re.search(r"(?m)^#{1,3}\s*(?:\d+\s*[.)、]?\s*){2,}", text):
        issues.append("标题前存在重复数字序号")

    brief_hint = _extract_brief_hint(text)
    if not brief_hint:
        issues.append("缺少“一句话心理提示（120字以内）”内容")
    elif _count_compact_chars(brief_hint) > 120:
        issues.append(f"一句话心理提示超长（{_count_compact_chars(brief_hint)}字）")

    if not _has_encouraging_tail(text):
        issues.append("结尾缺少单独鼓励语")

    return issues


def _insert_or_replace_brief_hint(report, brief_hint):
    content = str(report or "").rstrip()
    heading_pattern = re.compile(
        r"(?ms)(^#{2,3}\s*一句话心理提示[^\n]*\n)(.*?)(?=^\s*#{1,3}\s|\Z)"
    )
    match = heading_pattern.search(content)
    if match:
        return (
            content[:match.start()]
            + match.group(1)
            + brief_hint.strip()
            + "\n"
            + content[match.end():].lstrip("\n")
        ).rstrip()

    if content:
        content += "\n\n"
    return f"{content}### 一句话心理提示（120字以内）\n{brief_hint.strip()}".rstrip()


def _ensure_report_tail_requirements(report, game_data):
    content = _clean_report_markdown(report)
    brief_hint = _extract_brief_hint(content)
    if not brief_hint or _count_compact_chars(brief_hint) > 120:
        content = _insert_or_replace_brief_hint(content, _build_fallback_brief_hint(game_data))

    if not _has_encouraging_tail(content):
        content = content.rstrip() + "\n\n你已经在认真照顾自己，这一步很重要。"

    return _clean_report_markdown(content)


def _get_json_data():
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        raise PuzzleError("请求体必须是 JSON 对象")
    return data


def _safe_track(track_fn, *args, **kwargs):
    try:
        track_fn(*args, **kwargs)
    except Exception:
        app.logger.error("行为数据记录失败:\n%s", traceback.format_exc())


def _evaluate_report_level(report_text: str, user_questions: list = None) -> tuple:
    """评估心理报告等级，返回(等级, 备注)"""
    text = str(report_text or "").lower()

    # 如果有用户问题回答，也加入分析
    if user_questions:
        questions_text = " ".join([str(q) for q in user_questions if q]).lower()
        text = text + " " + questions_text

    # 关键词匹配规则
    severe_keywords = ["严重", "极度", "强烈焦虑", "抑郁倾向", "自我否定", "孤立", "逃避", "困扰明显",
                       "想死", "自杀", "轻生", "活着没意思", "绝望", "崩溃", "无法承受"]
    problem_keywords = ["压力较大", "焦虑", "紧张", "不安", "疲惫", "困惑", "犹豫", "负担",
                        "痛苦", "难受", "失眠", "烦躁", "迷茫", "无助", "孤独"]
    good_keywords = ["稳定", "良好", "积极", "平衡", "适应", "健康", "正常"]
    excellent_keywords = ["优秀", "出色", "自信", "乐观", "充满活力", "目标明确"]

    severe_count = sum(1 for kw in severe_keywords if kw in text)
    problem_count = sum(1 for kw in problem_keywords if kw in text)
    good_count = sum(1 for kw in good_keywords if kw in text)
    excellent_count = sum(1 for kw in excellent_keywords if kw in text)

    # 评级逻辑（负面情绪优先）
    if severe_count >= 1:
        matched_keywords = [kw for kw in severe_keywords if kw in text][:3]
        if severe_count >= 2:
            reason = f"检测到严重负面情绪或风险信号：{', '.join(matched_keywords)}"
            return "有大问题", reason
        else:
            reason = f"检测到负面心理指标：{', '.join(matched_keywords)}"
            return "有问题", reason
    elif problem_count >= 2:
        matched_keywords = [kw for kw in problem_keywords if kw in text][:3]
        reason = f"检测到负面心理指标：{', '.join(matched_keywords)}"
        return "有问题", reason
    elif problem_count >= 1:
        return "一般", "检测到轻微负面情绪"
    elif excellent_count >= 2 and good_count >= 2:
        return "优秀", "无"
    elif good_count >= 2:
        return "良好", "无"
    else:
        return "一般", "无"


def _generate_report_async(task_id: str, client_id: str, game_id: str, image_source: str, game_data: dict):
    """异步生成报告的后台任务"""
    try:
        with report_generation_lock:
            report_generation_tasks[task_id]["status"] = "processing"

        # 验证图片
        image_check = _validate_image_source_rule(image_source)
        if not image_check.get("valid"):
            with report_generation_lock:
                report_generation_tasks[task_id]["status"] = "failed"
                report_generation_tasks[task_id]["error"] = image_check.get("message", "图片校验失败")
            return

        # 同步会话数据
        _safe_track(
            analytics_store.update_from_report_payload,
            client_id,
            game_id,
            _compact_image_source(image_source, image_check),
            game_data
        )

        # 获取近期行为摘要
        recent_behavior_summary = "近期行为数据：暂无记录，仅基于本次拼图行为分析。"
        try:
            recent_behavior_summary = analytics_store.build_recent_behavior_prompt(client_id)
        except Exception:
            pass

        image_ref = _safe_image_prompt_ref(image_source, image_check)
        image_element_line = ""
        if image_check.get("isCustom"):
            el = image_check.get("elements", {})
            image_element_line = f"- 图片内容要素：房子={el.get('house')}, 树={el.get('tree')}, 人物={el.get('person')}\n"

        behavior_tags = _build_behavior_tags(game_data)
        behavior_tag_text = "\n".join(f"- {tag}" for tag in behavior_tags)

        user_message = f"""
用户场景：高校心理健康教育（大学生）
用户选择的图片：{image_ref}

游戏数据：
- 完成时间：{game_data.get('completionTime', '未知')}
- 移动步数：{game_data.get('moveCount', 0)}
- 难度等级：{game_data.get('difficulty', '未知')}
- 拼图顺序：{game_data.get('pieceOrder', '未记录')}
- 操作时间间隔：{game_data.get('timeIntervals', '未记录')}
- 修改次数：{game_data.get('modificationCount', 0)}
- 用户标识：{client_id}
{image_element_line}

行为标签（由系统计算）：
{behavior_tag_text}

近期行为数据摘要：
{recent_behavior_summary}

请基于"本次游戏数据 + 近期行为摘要"生成大学生心理分析报告。
"""

        # 调用AI生成报告
        api_client = get_client()
        response = api_client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_message}
            ],
            stream=False,
            temperature=0.7,
            max_tokens=2000
        )

        report = _clean_report_markdown(response.choices[0].message.content)
        issues = _collect_report_issues(report)
        if issues:
            fix_items = "\n".join(f"- {item}" for item in issues)
            try:
                rewrite_response = api_client.chat.completions.create(
                    model="deepseek-chat",
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": user_message},
                        {"role": "assistant", "content": report},
                        {"role": "user", "content": REWRITE_PROMPT + "\n\n" + fix_items},
                    ],
                    stream=False,
                    temperature=0.3,
                    max_tokens=2000
                )
                rewritten = rewrite_response.choices[0].message.content
                if rewritten:
                    report = _clean_report_markdown(rewritten)
            except Exception:
                pass

        report = _ensure_report_tail_requirements(report, game_data)

        # 保存报告
        _safe_track(
            analytics_store.save_report,
            client_id,
            game_id,
            _compact_image_source(image_source, image_check),
            user_message,
            report
        )

        # 更新任务状态
        with report_generation_lock:
            report_generation_tasks[task_id]["status"] = "completed"
            report_generation_tasks[task_id]["report"] = report
            report_generation_tasks[task_id]["timestamp"] = time.time()

    except Exception as e:
        with report_generation_lock:
            report_generation_tasks[task_id]["status"] = "failed"
            report_generation_tasks[task_id]["error"] = str(e)


@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查接口"""
    return jsonify({"status": "ok", "message": "Backend is running"}), 200


@app.route('/api/puzzle/games', methods=['POST'])
def create_puzzle_game():
    """创建拼图局"""
    try:
        data = _get_json_data()
        image_source = data.get("imageSource", "")
        grid_size = data.get("gridSize", 3)
        modifiers = data.get("modifiers", {})
        client_id = normalize_client_id(data.get("clientId"))

        # 调试日志
        app.logger.info(f"创建拼图游戏 - gridSize: {grid_size}, modifiers: {modifiers}")

        image_check = _validate_image_source_rule(image_source)
        if not image_check.get("valid"):
            return jsonify({
                "error": "图片校验未通过",
                "message": image_check.get("message", "上传图片不符合房树人三要素要求。"),
                "details": image_check,
            }), 403

        state = puzzle_engine.create_game(
            image_source=image_source,
            grid_size=grid_size,
            modifiers=modifiers
        )

        # 调试日志
        app.logger.info(f"拼图创建成功 - tray数量: {len(state.get('tray', []))}, hiddenCount: {state.get('hiddenCount', 0)}")

        _safe_track(
            analytics_store.upsert_game_session,
            client_id,
            state,
            image_source=_compact_image_source(image_source, image_check)
        )
        return jsonify({"success": True, "state": state, "imageCheck": image_check}), 200
    except PuzzleError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "创建拼图失败", "message": str(e)}), 500


@app.route('/api/puzzle/games/<game_id>', methods=['GET'])
def get_puzzle_game(game_id):
    """获取拼图局状态"""
    try:
        state = puzzle_engine.get_game_state(game_id)
        return jsonify({"success": True, "state": state}), 200
    except PuzzleError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "获取拼图状态失败", "message": str(e)}), 500


@app.route('/api/puzzle/games/<game_id>/actions', methods=['POST'])
def apply_puzzle_action(game_id):
    """执行拼图动作"""
    try:
        data = _get_json_data()
        action = data.get("action", "")
        payload = data.get("payload", {})
        if not isinstance(payload, dict):
            payload = {}
        client_id = normalize_client_id(data.get("clientId"))

        state = puzzle_engine.apply_action(game_id, action, payload)
        _safe_track(analytics_store.log_action, client_id, game_id, action, payload, state)
        _safe_track(
            analytics_store.upsert_game_session,
            client_id,
            state,
            image_source=_compact_image_source(state.get("imageSource", ""))
        )
        return jsonify({"success": True, "state": state}), 200
    except PuzzleError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "执行拼图动作失败", "message": str(e)}), 500

@app.route('/api/generate-report', methods=['POST'])
def generate_report():
    """生成心理分析报告"""
    try:
        # 1. 获取请求数据
        data = _get_json_data()
        client_id = normalize_client_id(data.get("clientId"))
        game_id = str(data.get("gameId", "") or "").strip()

        # 2. 验证图片来源
        image_source = data.get("imageSource", "")
        image_check = _validate_image_source_rule(image_source)
        if not image_check.get("valid"):
            return jsonify({
                "error": "图片校验未通过",
                "message": image_check.get("message", "上传图片不符合房树人三要素要求。"),
                "details": image_check,
            }), 403

        # 3. 获取游戏数据
        game_data = data.get("gameData", {})

        if not game_data:
            return jsonify({"error": "游戏数据为空"}), 400

        # 3.1 同步本次会话到近期数据库
        _safe_track(
            analytics_store.update_from_report_payload,
            client_id,
            game_id,
            _compact_image_source(image_source, image_check),
            game_data
        )

        # 3.2 读取近期行为摘要并作为补充证据
        recent_behavior_summary = "近期行为数据：暂无记录，仅基于本次拼图行为分析。"
        try:
            recent_behavior_summary = analytics_store.build_recent_behavior_prompt(client_id)
        except Exception:
            app.logger.error("读取近期行为摘要失败:\n%s", traceback.format_exc())

        image_ref = _safe_image_prompt_ref(image_source, image_check)
        image_element_line = ""
        if image_check.get("isCustom"):
            el = image_check.get("elements", {})
            image_element_line = f"- 图片内容要素：房子={el.get('house')}, 树={el.get('tree')}, 人物={el.get('person')}\n"

        behavior_tags = _build_behavior_tags(game_data)
        behavior_tag_text = "\n".join(f"- {tag}" for tag in behavior_tags)

        # 4. 构建用户消息
        user_message = f"""
用户场景：高校心理健康教育（大学生）
用户选择的图片：{image_ref}

游戏数据：
- 完成时间：{game_data.get('completionTime', '未知')}
- 移动步数：{game_data.get('moveCount', 0)}
- 难度等级：{game_data.get('difficulty', '未知')}
- 拼图顺序：{game_data.get('pieceOrder', '未记录')}
- 操作时间间隔：{game_data.get('timeIntervals', '未记录')}
- 修改次数：{game_data.get('modificationCount', 0)}
- 用户标识：{client_id}
{image_element_line}

行为标签（由系统计算）：
{behavior_tag_text}

近期行为数据摘要：
{recent_behavior_summary}

请基于“本次游戏数据 + 近期行为摘要”生成大学生心理分析报告。
请严格遵守：
1. 本次数据优先，近期摘要只作补充，不得过度推断；
2. 不得给出医学诊断结论；
3. 结论需绑定行为证据；
4. 标题不要数字序号；
5. 必须包含“一句话心理提示（120字以内）”和最后的单独鼓励语。
"""

        # 5. 调用DeepSeek API
        api_client = get_client()
        response = api_client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_message}
            ],
            stream=False,
            temperature=0.7,
            max_tokens=2000
        )

        report = _clean_report_markdown(response.choices[0].message.content)

        issues = _collect_report_issues(report)
        if issues:
            app.logger.warning("报告初稿存在合规问题，触发一次修订: %s", " | ".join(issues))
            fix_items = "\n".join(f"- {item}" for item in issues)
            rewrite_request = f"""
以下是报告初稿中需要修复的问题：
{fix_items}

请在不改变事实依据的前提下输出修订后的完整报告。
"""
            try:
                rewrite_response = api_client.chat.completions.create(
                    model="deepseek-chat",
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": user_message},
                        {"role": "assistant", "content": report},
                        {"role": "user", "content": REWRITE_PROMPT + "\n\n" + rewrite_request},
                    ],
                    stream=False,
                    temperature=0.3,
                    max_tokens=2000
                )
                rewritten = rewrite_response.choices[0].message.content
                if rewritten:
                    report = _clean_report_markdown(rewritten)
            except Exception:
                app.logger.error("报告自动修订失败:\n%s", traceback.format_exc())

        report = _ensure_report_tail_requirements(report, game_data)

        # 6. 返回结果
        _safe_track(
            analytics_store.save_report,
            client_id,
            game_id,
            _compact_image_source(image_source, image_check),
            user_message,
            report
        )
        return jsonify({
            "success": True,
            "report": report,
            "imageSource": image_source,
            "imageCheck": image_check,
            "timestamp": time.time()
        }), 200

    except PuzzleError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        app.logger.error(f"生成报告时出错: {str(e)}")
        return jsonify({
            "error": "生成报告失败",
            "message": str(e)
        }), 500

@app.route('/api/validate-image', methods=['POST'])
def validate_image():
    """验证图片是否可以被分析"""
    try:
        data = _get_json_data()
        image_source = data.get("imageSource", "")
        check = _validate_image_source_rule(image_source)
        return jsonify(check), 200
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({
            "error": "验证失败",
            "message": str(e)
        }), 500


@app.route('/api/reports/list', methods=['POST'])
def list_reports():
    """获取用户的报告列表"""
    try:
        data = _get_json_data()
        client_id = normalize_client_id(data.get("clientId"))
        reports = analytics_store.get_reports_by_client(client_id, limit=50)

        # 格式化返回数据
        result = []
        for report in reports:
            result.append({
                "id": report["id"],
                "gameId": report["game_id"],
                "imageSource": report["image_source"],
                "reportText": report["report_text"],
                "createdAt": report["created_at"],
                "createdAtFormatted": datetime.fromtimestamp(report["created_at"]).strftime("%Y-%m-%d %H:%M:%S")
            })

        return jsonify({"success": True, "reports": result}), 200
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "获取报告列表失败", "message": str(e)}), 500


@app.route('/api/reports/generate-async', methods=['POST'])
def generate_report_async():
    """异步生成报告"""
    try:
        data = _get_json_data()
        client_id = normalize_client_id(data.get("clientId"))
        game_id = str(data.get("gameId", "") or "").strip()
        image_source = data.get("imageSource", "")
        game_data = data.get("gameData", {})

        if not game_data:
            return jsonify({"error": "游戏数据为空"}), 400

        # 创建任务ID
        task_id = str(uuid.uuid4())

        # 初始化任务状态
        with report_generation_lock:
            report_generation_tasks[task_id] = {
                "status": "pending",
                "clientId": client_id,
                "createdAt": time.time()
            }

        # 启动后台线程生成报告
        thread = threading.Thread(
            target=_generate_report_async,
            args=(task_id, client_id, game_id, image_source, game_data)
        )
        thread.daemon = True
        thread.start()

        return jsonify({"success": True, "taskId": task_id}), 200
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "启动报告生成失败", "message": str(e)}), 500


@app.route('/api/reports/task-status/<task_id>', methods=['GET'])
def get_report_task_status(task_id):
    """查询报告生成任务状态"""
    try:
        with report_generation_lock:
            task = report_generation_tasks.get(task_id)

        if not task:
            return jsonify({"error": "任务不存在"}), 404

        result = {
            "taskId": task_id,
            "status": task["status"],
            "createdAt": task["createdAt"]
        }

        if task["status"] == "completed":
            result["report"] = task.get("report", "")
            result["timestamp"] = task.get("timestamp", time.time())
        elif task["status"] == "failed":
            result["error"] = task.get("error", "未知错误")

        return jsonify(result), 200
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "查询任务状态失败", "message": str(e)}), 500


@app.route('/api/healing/sessions/<int:report_id>', methods=['GET'])
def get_healing_sessions(report_id):
    """获取某个报告的所有疗愈会话"""
    try:
        sessions = analytics_store.get_healing_sessions_by_report(report_id)
        result = []
        for session in sessions:
            result.append({
                "sessionId": session["session_id"],
                "questionCount": session["question_count"],
                "isCompleted": session["question_count"] >= 3,
                "isDeleted": bool(session.get("is_deleted", 0)),
                "createdAt": session["created_at"],
                "updatedAt": session["updated_at"]
            })
        return jsonify({"success": True, "sessions": result}), 200
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "获取会话列表失败", "message": str(e)}), 500


@app.route('/api/healing/session/<session_id>', methods=['GET'])
def get_healing_session_detail(session_id):
    """获取会话详情和消息历史"""
    try:
        session = analytics_store.get_healing_session_by_id(session_id)
        if not session:
            return jsonify({"error": "会话不存在"}), 404

        messages = analytics_store.get_healing_messages(session_id)
        return jsonify({
            "success": True,
            "session": {
                "sessionId": session["session_id"],
                "questionCount": session["question_count"],
                "isCompleted": session["question_count"] >= 3
            },
            "messages": messages
        }), 200
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "获取会话详情失败", "message": str(e)}), 500


@app.route('/api/healing/create-session', methods=['POST'])
def create_healing_session():
    """创建心理疗愈会话"""
    try:
        data = _get_json_data()
        client_id = normalize_client_id(data.get("clientId"))
        report_id = int(data.get("reportId", 0))
        report_content = data.get("reportContent", "")

        session_id = str(uuid.uuid4())
        analytics_store.create_healing_session(session_id, client_id, report_id, report_content)

        # 添加系统初始消息
        analytics_store.add_healing_message(
            session_id,
            "system",
            f"已上传心理报告，报告内容：\n{report_content[:500]}..."
        )

        return jsonify({"success": True, "sessionId": session_id}), 200
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "创建疗愈会话失败", "message": str(e)}), 500


@app.route('/api/healing/chat', methods=['POST'])
def healing_chat():
    """心理疗愈对话"""
    try:
        data = _get_json_data()
        session_id = data.get("sessionId", "")
        user_message = data.get("message", "").strip()

        if not user_message:
            return jsonify({"error": "消息不能为空"}), 400

        # 检查问题数量限制
        question_count = analytics_store.get_question_count(session_id)
        if question_count >= 3:
            return jsonify({"error": "已达到提问次数上限（3次）"}), 403

        # 保存用户消息
        analytics_store.add_healing_message(session_id, "user", user_message)

        # 增加问题计数
        new_count = analytics_store.increment_question_count(session_id)

        # 获取对话历史
        messages = analytics_store.get_healing_messages(session_id)

        # 构建对话上下文
        chat_messages = [
            {
                "role": "system",
                "content": """你是一位温和的心理辅导助手，基于用户的心理报告提供支持性对话。

核心原则：
1. 用温暖、理解的语气回应，避免医学诊断
2. 重点在于倾听、共情和引导自我觉察
3. 每次回复控制在150字以内，简洁温暖
4. 通过开放式问题引导用户表达真实感受

风险识别：
- 如果用户表达自杀、自伤、绝望等严重负面情绪，要表达关切并建议寻求专业帮助
- 对于焦虑、压力、孤独等情绪，给予共情和支持性建议
- 鼓励用户分享具体情境和感受，而不是停留在表面

提问策略：
- 第一轮：了解用户当前最困扰的事情或感受
- 第二轮：深入探索具体情境和情绪体验
- 第三轮：引导思考应对方式和可能的支持资源"""
            }
        ]

        for msg in messages:
            if msg["role"] == "system":
                chat_messages.append({"role": "system", "content": msg["content"]})
            elif msg["role"] == "user":
                chat_messages.append({"role": "user", "content": msg["content"]})
            elif msg["role"] == "assistant":
                chat_messages.append({"role": "assistant", "content": msg["content"]})

        # 调用AI
        api_client = get_client()
        response = api_client.chat.completions.create(
            model="deepseek-chat",
            messages=chat_messages,
            stream=False,
            temperature=0.8,
            max_tokens=300
        )

        assistant_message = response.choices[0].message.content

        # 保存AI回复
        analytics_store.add_healing_message(session_id, "assistant", assistant_message)

        return jsonify({
            "success": True,
            "message": assistant_message,
            "questionCount": new_count,
            "remainingQuestions": 3 - new_count
        }), 200
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "对话失败", "message": str(e)}), 500


@app.route('/api/healing/submit-info', methods=['POST'])
def submit_healing_info():
    """提交用户信息"""
    try:
        data = _get_json_data()
        session_id = data.get("sessionId", "")
        user_name = data.get("userName", "").strip()
        user_student_id = data.get("userStudentId", "").strip()
        is_anonymous = data.get("isAnonymous", True)

        analytics_store.update_healing_user_info(session_id, user_name, user_student_id, is_anonymous)

        return jsonify({"success": True}), 200
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "提交信息失败", "message": str(e)}), 500


@app.route('/api/healing/delete-session', methods=['POST'])
def delete_healing_session():
    """软删除疗愈会话"""
    try:
        data = _get_json_data()
        session_id = data.get("sessionId", "")

        if not session_id:
            return jsonify({"error": "会话ID不能为空"}), 400

        analytics_store.soft_delete_healing_session(session_id)

        return jsonify({"success": True}), 200
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "删除会话失败", "message": str(e)}), 500


@app.route('/api/reports/delete', methods=['POST'])
def delete_report():
    """删除报告及其所有关联的疗愈会话"""
    try:
        data = _get_json_data()
        report_id = data.get("reportId", 0)

        if not report_id:
            return jsonify({"error": "报告ID不能为空"}), 400

        analytics_store.delete_report_and_sessions(report_id)

        return jsonify({"success": True}), 200
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "删除报告失败", "message": str(e)}), 500


@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """管理员登录验证"""
    try:
        data = _get_json_data()
        password = data.get("password", "")

        if password == ADMIN_PASSWORD:
            return jsonify({"success": True, "message": "登录成功"}), 200
        else:
            return jsonify({"success": False, "message": "密码错误"}), 401
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "登录失败", "message": str(e)}), 500


@app.route('/api/admin/export-data', methods=['GET'])
def admin_export_data():
    """管理员导出数据为Excel"""
    try:
        if Workbook is None:
            return jsonify({"error": "服务器缺少Excel导出组件"}), 500

        # 获取所有疗愈数据
        healing_data = analytics_store.get_all_healing_data()

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "心理测试数据"

        # 设置表头
        headers = ["ID", "测试时间", "姓名", "学号", "问题1", "问题2", "问题3", "心理报告结论", "备注"]
        ws.append(headers)

        # 设置表头样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 按日期分组并生成ID
        from collections import defaultdict
        date_groups = defaultdict(list)
        for session in healing_data:
            created_at = session.get("created_at", 0)
            date_str = datetime.fromtimestamp(created_at).strftime("%Y/%m/%d")
            date_groups[date_str].append(session)

        # 填充数据
        row_num = 2
        for date_str in sorted(date_groups.keys()):
            sessions = sorted(date_groups[date_str], key=lambda x: x.get("created_at", 0))
            for idx, session in enumerate(sessions, 1):
                # 生成ID: 日期_序号
                session_id = f"{date_str}_{idx:02d}"

                # 测试时间
                created_at = session.get("created_at", 0)
                test_time = datetime.fromtimestamp(created_at).strftime("%Y-%m-%d %H:%M:%S")

                user_name = session.get("user_name", "") if not session.get("is_anonymous") else "匿名"
                user_student_id = session.get("user_student_id", "") if not session.get("is_anonymous") else "匿名"

                questions = session.get("questions", [])
                q1 = questions[0] if len(questions) > 0 else ""
                q2 = questions[1] if len(questions) > 1 else ""
                q3 = questions[2] if len(questions) > 2 else ""

                # 评估报告等级（同时分析报告和用户回答）
                report_content = session.get("report_content", "")
                level, reason = _evaluate_report_level(report_content, questions)

                row_data = [session_id, test_time, user_name, user_student_id, q1, q2, q3, level, reason]
                ws.append(row_data)

                # 如果是"有问题"或"有大问题"，标红
                if level in ["有问题", "有大问题"]:
                    level_cell = ws.cell(row=row_num, column=8)
                    level_cell.font = Font(color="FF0000", bold=True)

                row_num += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 40

        # 保存到临时文件
        temp_file = os.path.join(os.path.dirname(__file__), "data", f"export_{int(time.time())}.xlsx")
        os.makedirs(os.path.dirname(temp_file), exist_ok=True)
        wb.save(temp_file)

        # 发送文件
        return send_file(
            temp_file,
            as_attachment=True,
            download_name=f"心理测试数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"error": "导出数据失败", "message": str(e)}), 500



@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_frontend(path: str):
    """服务前端静态文件（生产环境单服务部署）。"""
    # /api/* 已由上方路由匹配，这里仅处理静态资源与前端路由回退
    if path:
        target = os.path.join(FRONTEND_DIR, path)
        if os.path.isfile(target):
            return send_from_directory(FRONTEND_DIR, path)
        # 缺失的静态资源直接返回 404，避免把 HTML 当成 JS/CSS 解析
        if "." in os.path.basename(path):
            return jsonify({"error": "not found"}), 404
    resp = make_response(send_from_directory(FRONTEND_DIR, "index.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

if __name__ == '__main__':
    # 检查API密钥
    if not os.environ.get("DEEPSEEK_API_KEY"):
        print("警告: 未设置 DEEPSEEK_API_KEY 环境变量")

    # 启动服务器
    app.run(host='0.0.0.0', port=5000, debug=True)
